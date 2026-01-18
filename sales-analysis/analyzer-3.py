# analyzer-3.py
import os
import pandas as pd
from helpers import calculate_total, format_currency
import openpyxl
from pypdf import PdfWriter
from fpdf import FPDF


# Read data
df = pd.read_csv('data/sales.csv')

# Calculate total for each row
totals = []
for index, row in df.iterrows():
    total = calculate_total(row['quantity'], row['price'])
    totals.append(total)

# Add totals to our data
df['total'] = totals

# Display with formatted totals
print("Sales Data:")
for index, row in df.iterrows():
    formatted_total = format_currency(row['total'])
    print(f"{row['product']}: {formatted_total}")

# Show grand total
grand_total = df['total'].sum()
formatted_grand_total = format_currency(grand_total)
print(f"\nGrand Total: {formatted_grand_total}")

# Create output directory
os.makedirs('output', exist_ok=True)

# Save as different formats
# 1. JSON format (good for web APIs)
df.to_json('output/sales_data.json', orient='records', indent=2)

# 2. Excel format (good for sharing)
df.to_excel('output/sales_data.xlsx', index=False)

# 3. Updated CSV (with our new total column)
df.to_csv('output/sales_with_totals.csv', index=False)

# 4. PDF format
pdf = FPDF()
pdf.add_page()
pdf.set_font("Arial", "B", 16)  
pdf.cell(0, 10, "Sales Report", ln=True, align="C")
pdf.set_font("Arial", size=12)
pdf.ln(5)

# Add table header
pdf.set_font("Arial", "B", 10)
pdf.cell(60, 10, "Product", border=1)
pdf.cell(40, 10, "Quantity", border=1)
pdf.cell(40, 10, "Price", border=1) 
pdf.cell(50, 10, "Total", border=1, ln=True)

# Add data rows
pdf.set_font("Arial", size=10)
for index, row in df.iterrows():
    pdf.cell(60, 10, str(row['product']), border=1)
    pdf.cell(40, 10, str(row['quantity']), border=1)
    pdf.cell(40, 10, format_currency(row['price']), border=1)
    pdf.cell(50, 10, format_currency(row['total']), border=1, ln=True)

# Add grand total
pdf.set_font("Arial", "B", 10)
pdf.cell(140, 10, "Grand Total:", border=0, align="R")
pdf.cell(50, 10, format_currency(grand_total), border=1, ln=True)

pdf.output('output/sales_data.pdf')

print("\nFiles saved:")
print("- output/sales_data.json")
print("- output/sales_data.xlsx") 
print("- output/sales_with_totals.csv")
print("- output/sales_data.pdf")